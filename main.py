

from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.remote.webelement import By
from time import sleep
#  for excel
import openpyxl as O
Excel_file = "vietstockSave.xlsx"
wb = O.load_workbook(Excel_file)
pageError = 21
pageName  = "Dịch vụ lưu trú, ăn uống"
# .table > tbody:nth-child(2) > tr:nth-child(1) > td:nth-child(2) > a:nth-child(1)
WEB_URL = "https://finance.vietstock.vn/chi-so-nganh.htm"
driver = webdriver.Chrome()
driver.get(WEB_URL)
# // Get ROW COUNT
colCount = len(driver.find_elements_by_xpath("/html/body/div[2]/div[11]/div/div[1]/table/tbody/tr[1]/td"))
# Get Column Count
rowCount = len(driver.find_elements_by_xpath("/html/body/div[2]/div[11]/div/div[1]/table/tbody/tr"))
#  excel setup
# print("Row count :" + rowCount)
print(colCount)
print(rowCount)
# // process data
# Printing the data of the table
for r in range(1, rowCount +1): # rowCount +1
    row_num =1
    col_num =1
    ws =0
    for p in range(2, colCount +1):
        # obtaining the text from each column of the table
        value = driver.find_element_by_xpath(
            "/html/body/div[2]/div[11]/div/div[1]/table/tbody/tr[" + str(
                r) + "]/td[" + str(p) + "]").text
        print(value, end='       ')
        if p == 2:
            if r == pageError:
                ws = wb[pageName]
            else:
                ws = wb[value]
            row_num = ws.max_row
            col_num = 1
        else:
            ws.cell((row_num + 1), col_num).value = value;
            col_num += 1
    print()
wb.save(Excel_file)
wb.close()
driver.close()
# excel
#
# import openpyxl as O
# Excel_file = "test1.xlsx"
# Excel_worksheet = "Bán buôn"
#
# wb = O.load_workbook(Excel_file)
# ws = wb[Excel_worksheet]
# row_num = ws.max_row
# col_num = ws.max_column
#
# print ("The number of rows is ", row_num, "and the number of columns is ", col_num)